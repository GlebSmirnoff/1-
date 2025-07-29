from django.core.management.base import BaseCommand
import openpyxl
from listing.models import Listing
from django.contrib.auth import get_user_model

User = get_user_model()

REQUIRED_COLUMNS = [
    "Марка", "Модель", "Рік", "Кузов", "Двигун", "Кількість власників"
]

class Command(BaseCommand):
    help = "Імпорт оголошень з Excel файлу"

    def add_arguments(self, parser):
        parser.add_argument("filepath", type=str, help="Шлях до Excel-файлу")

    def handle(self, *args, **options):
        filepath = options["filepath"]
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        # Отримуємо заголовки (1-й рядок)
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip().title() if cell else "" for cell in raw_headers]
        self.stdout.write(f"🔍 Заголовки з файлу (нормалізовано): {headers}")

        # Перевірка наявності всіх обов’язкових колонок
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            self.stdout.write(self.style.ERROR(f"❌ Відсутні обов'язкові колонки: {', '.join(missing)}"))
            return

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        imported, skipped = 0, 0

        for row in rows:
            data = dict(zip(headers, [str(cell).strip() if cell else "" for cell in row]))

            if not all(data.get(col) for col in REQUIRED_COLUMNS):
                self.stdout.write(self.style.WARNING(f"❌ Пропущено рядок: недостатньо даних"))
                skipped += 1
                continue

            try:
                user = User.objects.filter(is_superuser=True).first()
                if not user:
                    self.stdout.write(self.style.ERROR("❌ Не знайдено суперадміна для імпорту"))
                    break

                Listing.objects.create(
                    user=user,
                    vin=None,
                    color=None,
                    owners_count=int(data["Кількість Власників"]),
                    price=0,
                    currency="USD",
                    year=int(data["Рік"]),
                    mileage=0,
                    brand=None,
                    model=None,
                    body_type=None,
                    engine=None,
                )
                imported += 1
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Помилка: {e}"))
                skipped += 1

        self.stdout.write(self.style.SUCCESS(f"✅ Імпорт завершено: {imported} додано, {skipped} пропущено"))
